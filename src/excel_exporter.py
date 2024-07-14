import logging
import os
import subprocess
from PyQt6.QtWidgets import QFileDialog, QMessageBox, QPushButton, QVBoxLayout, QWidget
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class ExcelExporter:
    @staticmethod
    def export_to_excel(parent, company_table, collection):
        try:
            filename, _ = QFileDialog.getSaveFileName(parent, "Export Excel", "", "Excel Files (*.xlsx)")
            if not filename:
                return

            if not filename.endswith('.xlsx'):
                filename += '.xlsx'

            if collection == "Company_Install":
                headers = [
                    "Telephely név", "Összes terminál igény", "Kiadva", "Telepítés",
                    "Áram", "Elosztó", "Hálózat", "PTG", "Szoftver", "Param", "Helyszín",
                    "Teszt", "Véglegesítve", "Megjegyzés",
                    "Megjegyzés ideje", "Véglegesítés ideje"
                ]
            else:  # Company_Demolition
                headers = [
                    "Telephely név", "Összes terminál igény", "Bontás", "Felszerelés", "Bázis Leszerelés",
                    "Megjegyzés", "Megjegyzés ideje", "Véglegesítés ideje"
                ]

            wb = Workbook()
            ws = wb.active

            # Add headers with formatting
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add data
            for row in range(company_table.rowCount()):
                row_data = []
                for col in range(1, company_table.columnCount()):
                    item = company_table.item(row, col)
                    value = item.text() if item else ""
                    if value.lower() in ['true', 'false', 'van', 'nincs']:
                        value = "Van" if value.lower() in ['true', 'van'] else "Nincs"
                    row_data.append(value)

                if collection == "Company_Install":
                    mapped_data = ExcelExporter.map_install_data(row_data)
                else:  # Company_Demolition
                    mapped_data = ExcelExporter.map_demolition_data(row_data)

                ws.append(mapped_data)

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(filename)

            # Create a custom message box with an "Open File" button
            msg_box = QMessageBox(parent)
            msg_box.setWindowTitle("Export Complete")
            msg_box.setText(f"Data exported to {filename}")

            open_button = QPushButton("Open File")
            open_button.clicked.connect(lambda: ExcelExporter.open_file(filename))

            layout = QVBoxLayout()
            layout.addWidget(open_button)

            widget = QWidget()
            widget.setLayout(layout)

            msg_box.layout().addWidget(widget, 1, 1)
            msg_box.exec()

        except Exception as e:
            logging.error(f"Error exporting to Excel: {e}")
            QMessageBox.critical(parent, "Error", f"Failed to export to Excel: {str(e)}")

    @staticmethod
    def map_install_data(row_data):
        return [
            row_data[1],  # Telephely név (CompanyName)
            row_data[3],  # Összes terminál igény (Quantity)
            "Van" if row_data[5] == "KIADVA" else "Nincs",  # Kiadva
            row_data[5],  # Telepítés
            row_data[6],  # Áram
            row_data[7],  # Elosztó
            row_data[8],  # Hálózat
            row_data[9],  # PTG
            row_data[10],  # Szoftver
            row_data[11],  # Param
            row_data[12],  # Helyszín
            "Van" if row_data[5] == "HELYSZINEN_TESZTELVE" else "Nincs",  # Teszt
            "Van" if row_data[5] == "KIRAKVA" else "Nincs",  # Véglegesítve
            "",  # Megjegyzés (not available)
            "",  # Megjegyzés ideje (not available)
            row_data[-1]  # Véglegesítés ideje (LastModified)
        ]

    @staticmethod
    def map_demolition_data(row_data):
        return [
            row_data[1],  # Telephely név (CompanyName)
            row_data[3],  # Összes terminál igény (Quantity)
            row_data[5],  # Bontás
            row_data[6],  # Felszerelés
            "Van" if row_data[7] == "True" else "Nincs",  # Bázis Leszerelés
            "",  # Megjegyzés (not available)
            "",  # Megjegyzés ideje (not available)
            row_data[-1]  # Véglegesítés ideje (LastModified)
        ]

    @staticmethod
    def open_file(filename):
        if os.path.exists(filename):
            if os.name == 'nt':  # For Windows
                os.startfile(filename)
            elif os.name == 'posix':  # For macOS and Linux
                subprocess.call(('open', filename))
            else:
                logging.error("Unsupported operating system")
        else:
            logging.error(f"File not found: {filename}")